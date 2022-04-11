from ecommerce_blueprint_version import db
from ecommerce_blueprint_version.models import Order, OrderedProduct, Product, WeChatOrder

from openpyxl import load_workbook

def convert_float_to_int(float_number):
   return int(float_number) if float_number and isinstance(float_number, float) else float_number

def import_excel(excel):
   wb = load_workbook(excel)
   sheet = wb[wb.sheetnames[-2]]
   return sheet.iter_rows(min_row=2, min_col=2, values_only=True)

def get_orders(current_user):
   orders = Order.query.join(OrderedProduct, Order.orderid == OrderedProduct.orderid) \
        .join(Product, OrderedProduct.productid == Product.productid) \
        .with_entities(Order.orderid, Order.order_date, Order.total_price, Order.address, Order.shipping_fee, Order.note,
                        Product.productid, Product.product_name_zh, Product.product_name_en, Product.product_name_fr,
                        OrderedProduct.variant_name_zh, OrderedProduct.variant_name_en, OrderedProduct.variant_name_fr, OrderedProduct.price,
                        OrderedProduct.quantity, OrderedProduct.sum
                        ) \
        .filter(Order.userid == current_user.userid) \
        .order_by(Order.orderid.desc()) \
        .all()

   return orders

# Gets form data for the sales transaction
def place_order(current_user, orderData, totalPrice, address, note, shipping_fee):

   # Since using with_entities, the result will be a tuple, we need to get the first element by using index 0
   userId, username = current_user.userid, current_user.username
   print('order data', address, note)
   # This part simply create an order with order date, total_price and userId
   order = Order(total_price=totalPrice, userid=userId, address=address, note=note, shipping_fee=shipping_fee)
   db.session.add(order)
   db.session.flush()
   db.session.commit()
   # Get the orderid 
   orderid = Order.query.with_entities(Order.orderid).filter(Order.userid == userId).order_by(
      Order.orderid.desc()).first()
   orderid = orderid[0]
   print("orderid", orderid)

   # add details to ordered_products table
   addOrderedproducts(orderid, orderData)
   
   # remove ordered products from cart after transaction is successful
   # sendtextconfirmation(phone,fullname,orderid)
   return (username, orderid)


def cancel_order(order_id):
   Order.query.filter(Order.orderid == order_id).delete()
   db.session.commit()

# adds data to orderdproduct table

def addOrderedproducts(orderid, orderData):
    # Since the cart data is saved in the localstorage, this function will not be used in this version
    # cart = Cart.query.with_entities(
    #     Cart.productid, Cart.quantity).filter(Cart.userid == userId)

    # Need to add orderproduct price and variant columns. Since we treat variants as different product, 
    # so two variants of the same product will run OrderedProduct twice
   for item in orderData:
      orderedproduct = OrderedProduct(
         orderid=orderid, productid=item.get("productid"), variant_name_zh=item.get("variant_name_zh"), variant_name_en=item.get("variant_name_en")
         , variant_name_fr=item.get("variant_name_fr"), price=item.get("price"), quantity=item.get("quantity"), sum=item.get("sum"))
      db.session.add(orderedproduct)
      db.session.flush()
      db.session.commit()